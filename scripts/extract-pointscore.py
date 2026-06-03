#!/usr/bin/env python3
"""Extract WWSC pointscore sheets from bryan-excel-original.xlsm.

Dumps the relevant pointscore sheets cell-by-cell (with formulas where present)
so we can document the working scoring source per Balerion's M3 directive.
Output is written as JSON for the rule artifact + a readable text summary.
"""
import json
import sys
import openpyxl

XLSM = "/Users/dino/Library/CloudStorage/Dropbox/Dino-Balerion-Claude-Code/Projekte/0004_swimming-app/bryan-excel-original.xlsm"

POINTSCORE_SHEETS = [
    "25m Point score", "50m Point score", "75m Point score",
    "Relay Point score",
    "Backstroke Pointscore", "Breaststroke Pointscore", "Butterfly Pointscore",
    "25m Brace Pointscore", "50m Brace Pointscore", "Medley Relay Pointscore",
    "Total Pointscore",
    "Event Times & Pointscores",
]

def dump_sheet(ws, max_rows=40, max_cols=20):
    rows = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row_cells = []
        any_val = False
        for c in range(1, min(ws.max_column, max_cols) + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                any_val = True
            row_cells.append(v)
        if any_val:
            rows.append({"row": r, "cells": row_cells})
    return {"max_row": ws.max_row, "max_col": ws.max_column, "rows": rows}

def main():
    print("Loading workbook (data_only=False to see formulas)...")
    wb_f = openpyxl.load_workbook(XLSM, data_only=False, read_only=True, keep_vba=False)
    print("Loading workbook (data_only=True to see computed values)...")
    wb_v = openpyxl.load_workbook(XLSM, data_only=True, read_only=True, keep_vba=False)

    result = {}
    for name in POINTSCORE_SHEETS:
        if name not in wb_f.sheetnames:
            print(f"  MISSING: {name}")
            continue
        print(f"  Extracting: {name}")
        ws_f = wb_f[name]
        ws_v = wb_v[name]
        result[name] = {
            "formulas": dump_sheet(ws_f),
            "values": dump_sheet(ws_v),
        }

    out = sys.argv[1] if len(sys.argv) > 1 else "/tmp/wwsc-pointscore-extract.json"
    with open(out, "w") as f:
        json.dump(result, f, indent=2, default=str)
    print(f"Wrote {out}")

    # Readable summary of the smallest pointscore sheet (likely the rule table)
    for name in ["25m Point score", "Total Pointscore"]:
        if name in result:
            print(f"\n===== {name} (values) =====")
            for row in result[name]["values"]["rows"][:25]:
                cells = [("" if c is None else str(c)) for c in row["cells"][:12]]
                print(f"  r{row['row']:>3}: " + " | ".join(cells))

if __name__ == "__main__":
    main()
